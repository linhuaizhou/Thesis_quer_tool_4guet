"""That's download_from_scihub in multy mode  by 1hz  """
import os

import openpyxl

from article_download.scihub import SciHub
from article_download.pdf2word import pdf2word as p2w
from win32com import client as wc


class download_from_scihub:
    def __init__(self, excel_input_path, sheet_name, pdf_out_path, word_out_path, txt_out_path, row_start,
                 column_start):
        self.excel_input_path = excel_input_path
        self.sheet_name = sheet_name
        self.pdf_out_path = pdf_out_path
        self.word_out_path = word_out_path
        self.txt_out_path = txt_out_path
        self.row_start = row_start
        self.column_start = column_start

    def Translate(self, input, output):
        # 转换
        wordapp = wc.Dispatch('kwps.Application')

        doc = wordapp.Documents.Open(input)
        # 为了让python可以在后续操作中r方式读取txt和不产生乱码，参数为4
        doc.SaveAs(output, 4)
        doc.Close()

    def do(self):
        doi_list = []
        wb = openpyxl.load_workbook(self.excel_input_path)
        sheet1 = wb[self.sheet_name]
        row_num = sheet1.max_row
        for row in range(self.row_start, row_num + 1):
            cell = sheet1.cell(row, self.column_start)
            if cell.value is not None:
                doi_list.append(cell.value)  # 读取EXCEL 1列的内容
        # print(doi_list)
        sh = SciHub()
        count = 0
        for i in doi_list:
            url1 = "https://sci-hub.se/" + str(i)
            # print(url1)
            #   print(type(url1))
            path1 = self.pdf_out_path + "/" + str(count) + ".pdf"
            # print(path1)
            # print(type(path1))
            result = sh.download(url1, path=path1)
            count += 1
            # print(str(count) + "finish!")
        print("PDF part has finished!Here are " + str(count) + "pdf file at location" + self.pdf_out_path)
        pdf_path_all_list = []
        for root, dirs, files in os.walk(self.pdf_out_path):
            for file in files:
                if os.path.splitext(file)[1] == '.pdf':
                    pdf_path_all_list.append(
                        os.path.join(root, file))  # 找到指定文件夹下pdf文件并将链接写进pdf_path_all_list
        count_1 = 0
        for i in pdf_path_all_list:
            pdf_path = i
            doc_path = self.word_out_path + "/" + str(count_1) + ".docx"
            p1 = p2w(pdf_path, doc_path)
            p2 = p1.do()
            input_file = doc_path
            output_file = self.txt_out_path + "/" + str(count_1) + ".txt"
            self.Translate(input_file, output_file)
            count_1 += 1
            print(str(count_1) + " finish!")
        print(
            "Word part and TXT part has finished!Check at location " + self.word_out_path + " and " + self.txt_out_path)
        print("DUO&PDF&WORD&TXT has finished!")

        # print("down")

