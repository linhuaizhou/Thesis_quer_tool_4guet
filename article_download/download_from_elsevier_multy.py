"""That's download_from_elsevier in multy mode  by 1hz  """
import openpyxl
import requests


class download_from_elsevier:
    def __init__(self, excel_input_path, sheet_name, txt_out_path):
        self.excel_input_path = excel_input_path
        self.sheet_name = sheet_name
        self.txt_out_path = txt_out_path

    def do(self):
        header = {'Accept': 'text/plain', 'CR-TDM-Rate-Limit': '4000', 'CR-TDM-Rate-Limit-Remaining': '76',
                  'CR-TDM-Rate-Limit-Reset': '1378072800'}

        url_publisher = "https://api.elsevier.com/content/article/doi/"  # 如果获取全文的话，将abstract替换成article
        APIKey = "爱思唯尔开发者API"  # 爱思唯尔开发者API
        arformat = "text/plain"  # text/xml,text/plain
        doi_list = []
        wb = openpyxl.load_workbook(self.excel_input_path)
        sheet1 = wb[self.sheet_name]
        row_num = sheet1.max_row
        count_3 = 0
        for row in range(2, row_num + 1):
            cell = sheet1.cell(row, 2)
            if cell.value is not None:
                count_3 += 1
                doi_list.append(cell.value)  # 读取EXCEL 1列的内容
        count_2 = 0
        print("There are " + str(count_3) + " DOI in excel.Starting  TXT&DUO downloading!")
        for i in doi_list:
            url = url_publisher + i + "?" + APIKey + "&httpAccept=" + arformat
            r = requests.get(url, headers=header)
            content = r.content.decode()
            path = self.txt_out_path + '/' + str(count_2) + '.txt'
            # print(path)
            f = open(path, 'w', encoding='utf-8')
            f.write(content)
            f.close()
            # self.signal.emit(str(count_2) + " has finished!")
            count_2 += 1
        # time.sleep(4)
        if count_2 == count_3:
            print("Duo&TXT has finish!Check txt at location " + str(self.txt_out_path))
