"""That's cha_bd_title_list_2_doi_list.py by 1hz  """
import time

import requests
from bs4 import BeautifulSoup
import urllib
import openpyxl


# 将Excel内的论文名称依次读取，爬取其百度学术返回的DOI，批量返回Excel
class title2doi:
    def __init__(self, title_input_excel, row_title_input, column_title_input, row_doi_output, column_doi_output):
        self.title_input_excel = title_input_excel
        self.column_title_input = column_title_input
        self.column_doi_output = column_doi_output
        self.row_title_input = row_title_input
        self.row_doi_output = row_doi_output

    def read_line(self, sheet_name, row_start, column_start, list_output):
        row_num = sheet_name.max_row
        for row in range(row_start, row_num + 1):
            cell = sheet_name.cell(row, column_start)
            if cell.value is not None:
                list_output.append(cell.value)  # 读取EXCEL 1列的内容

    def do(self):

        wb = openpyxl.load_workbook(self.title_input_excel)  # 注意这里是test
        sheet1 = wb["Sheet1"]
        list_title_input = []
        list_doi = []
        count_search = 1
        list_doing = []
        self.read_line(sheet1, self.row_title_input, self.column_title_input, list_title_input)

        print(list_title_input)

        for search in list_title_input:
            flag = False
            kwen = search.encode('utf-8')
            headers = {'Accept': 'text/plain', 'CR-TDM-Rate-Limit': '4000', 'CR-TDM-Rate-Limit-Remaining': '76',
                       'CR-TDM-Rate-Limit-Reset': '1378072800'}
            for x in range(1):  # 想要爬取多少页数据
                url = 'http://xueshu.baidu.com/s?wd=' + urllib.request.quote(kwen) + '&pn=' + str(
                    x * 10) + '&tn=SE_baiduxueshu_c1gjeupa&ie=utf-8&sc_f_para=sc_tasktype%3D%7BfirstSimpleSearch%7D&sc_hit=1'
                res = requests.get(url, headers=headers)
                bs1 = BeautifulSoup(res.text, 'html.parser')  # 解析数据
                list_titles = bs1.find_all('div', class_="sc_content")
                try:
                    count_1 = 0
                    for i in list_titles:
                        title = i.find('h3', class_="t c_font").text  # 爬到标题
                        half_link = i.find('h3', class_="t c_font").find('a')['href']
                        wholelink = str(half_link)
                        re = requests.get(wholelink, headers=headers)  # 爬取该网站内容
                        if re.status_code == 200:
                            bs2 = BeautifulSoup(re.text, 'html.parser')  # 解析该网站内容
                            infos = bs2.find('div', class_="main-info").find('div', class_="c_content")
                            papers = bs2.find('div', class_="paper_src_content")
                            if infos.find('div', class_="doi_wr") is not None:
                                doi = infos.find('div', class_="doi_wr").find('p', class_="kw_main").text.strip()

                                if count_1 == 0:
                                    list_doi.append(doi)
                                    list_doing.append(doi)
                                    print(count_search)
                                    print(doi)
                                    count_1 += 1
                                    flag = True
                                    count_search += 1

                                    break
                            else:
                                print('该文章无DOI, 详情请查看官网：' + wholelink + '\n')
                                print(count_search)
                        else:
                            print('该文章无链接')
                            count_search += 1
                            print(count_search)

                except:

                    print("error")
            if not flag:
                list_doi.append("")
        try:
            print(list_doi)
            count = self.row_doi_output
            for i in list_doi:
                sheet1.cell(row=count, column=self.column_doi_output, value=i)
                count += 1
            wb.save(self.title_input_excel)  # 注意这里是test
        except:
            print("error")
