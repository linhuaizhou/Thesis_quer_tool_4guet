"""That's cha_bd_mainword_search_excel.py by 1hz  """
import requests
from bs4 import BeautifulSoup
import urllib
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


# 该类是对百度学术指定关键词返回内容做excel输出,保存到指定的excel文档并输出，search为待搜索关键词，page_number为可搜索的页数，一般一页10个，input_excel为输出的excel路径#

class bd_main_word_search:
    def __init__(self, search, page_number, input_excel):
        self.search = search
        self.page_number = page_number
        self.input_excel = input_excel
        self.list_titles = []
        self.list_wholelink = []
        self.list_author = []
        self.list_abstract = []
        self.list_keywords = []
        self.list_year = []
        self.list_doi = []
        self.list_ref = []

    def do_search(self):
        kwen = self.search.encode('utf-8')  # 将汉字，用utf格式编码，赋值给gbkkw

        # 添加请求头，模拟浏览器正常访问，避免被反爬虫
        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36'}
        for x in range(self.page_number):  # 想要爬取多少页数据
            url = 'http://xueshu.baidu.com/s?wd=' + urllib.request.quote(kwen) + '&pn=' + str(
                x * 10) + '&tn=SE_baiduxueshu_c1gjeupa&ie=utf-8&sc_f_para=sc_tasktype%3D%7BfirstSimpleSearch%7D&sc_hit=1'
            res = requests.get(url, headers=headers)

            print(res.status_code)  # 查看是否能获取数据
            bs1 = BeautifulSoup(res.text, 'html.parser')  # 解析数据
            list_titles = bs1.find_all('div', class_="sc_content")
            for i in list_titles:
                title = i.find('h3', class_="t c_font").text  # 爬到标题
                print(title)
                self.list_titles.append(title)
                half_link = i.find('h3', class_="t c_font").find('a')['href']
                wholelink = str(half_link)
                print(wholelink)
                self.list_wholelink.append(wholelink)
                re = requests.get(wholelink, headers=headers)  # 爬取该网站内容
                if re.status_code == 200:
                    bs2 = BeautifulSoup(re.text, 'html.parser')  # 解析该网站内容
                    infos = bs2.find('div', class_="main-info").find('div', class_="c_content")
                    papers = bs2.find('div', class_="paper_src_content")
                    try:
                        if infos.find("div", class_="common_wr") is not None:  # 测试是否为专利
                            self.list_author.append("专利，无此内容")
                            self.list_abstract.append("专利，无此内容")
                            self.list_keywords.append("专利，无此内容")
                            self.list_year.append("专利，无此内容")
                            self.list_doi.append("专利，无此内容")
                            self.list_ref.append("专利，无此内容")
                        if infos.find('div', class_="author_wr") is not None:
                            author = infos.find('div', class_="author_wr").find('p', class_="author_text").text.strip()
                            print(author)
                            print(type(author))
                            self.list_author.append(author)
                        if infos.find('div', class_="author_wr") is None:
                            self.list_author.append(" ")
                        if infos.find('div', class_="abstract_wr") is not None:
                            abstract = infos.find('div', class_="abstract_wr").find('p', class_="abstract").text.strip()
                            print(abstract)
                            self.list_abstract.append(abstract)
                        if infos.find('div', class_="abstract_wr") is None:
                            self.list_abstract.append(" ")
                        if infos.find('div', class_="kw_wr") is not None:
                            keywords = infos.find('div', class_="kw_wr").find('p', class_="kw_main").text.strip()
                            print(keywords)
                            self.list_keywords.append(keywords)
                        if infos.find('div', class_="kw_wr") is None:
                            self.list_keywords.append(" ")
                        if infos.find('div', class_="year_wr") is not None:
                            year = infos.find('div', class_="year_wr").find('p', class_="kw_main").text.strip()
                            print(year)
                            self.list_year.append(year)
                        if infos.find('div', class_="year_wr") is None:
                            self.list_year.append(" ")
                        if infos.find('div', class_="doi_wr") is not None:
                            doi = infos.find('div', class_="doi_wr").find('p', class_="kw_main").text.strip()
                            print(doi)
                            self.list_doi.append(doi)
                        if infos.find('div', class_="doi_wr") is None:
                            self.list_doi.append(" ")
                        if infos.find('div', class_="ref_wr") is not None:
                            ref = infos.find('div', class_="ref_wr").find('p', class_="ref-wr-num").text.strip()
                            print(ref)
                            self.list_ref.append(ref)
                        if infos.find('div', class_="ref_wr") is None:
                            self.list_ref.append(" ")

                        else:
                            print('该文章无该内容, 详情请查看官网：' + wholelink + '\n')

                    except:
                        print("error!" + '\n')


                else:
                    print('该文章无链接')
                    self.list_author.append(" ")
                    self.list_abstract.append(" ")
                    self.list_keywords.append(" ")
                    self.list_year.append(" ")
                    self.list_doi.append(" ")
                    self.list_ref.append(" ")
        print(len(self.list_titles))
        print(len(self.list_wholelink))
        print(len(self.list_author))
        print(len(self.list_abstract))
        print(len(self.list_keywords))
        print(len(self.list_year))
        print(len(self.list_doi))
        print(len(self.list_ref))
        print(self.list_titles)
        print(self.list_wholelink)
        print(self.list_author)
        print(self.list_abstract)
        print(self.list_keywords)
        print(self.list_year)
        print(self.list_doi)
        print(self.list_ref)
        wb = openpyxl.Workbook(self.input_excel)
        wb.save(self.input_excel)
        wb = openpyxl.load_workbook(self.input_excel)  # 注意这里是test

        sheet1 = wb["Sheet"]
        try:
            count = 1
            for a, b, c, d, e, f, g, h in zip(self.list_titles, self.list_wholelink, self.list_author,
                                              self.list_abstract,
                                              self.list_keywords, self.list_year, self.list_doi, self.list_ref):
                sheet1.cell(row=count, column=1, value=count)
                sheet1.cell(row=count, column=2, value=ILLEGAL_CHARACTERS_RE.sub(r'', a))
                sheet1.cell(row=count, column=3, value=ILLEGAL_CHARACTERS_RE.sub(r'', b))
                sheet1.cell(row=count, column=4, value=ILLEGAL_CHARACTERS_RE.sub(r'', c))  # 防止乱码
                sheet1.cell(row=count, column=5, value=ILLEGAL_CHARACTERS_RE.sub(r'', d))
                sheet1.cell(row=count, column=6, value=ILLEGAL_CHARACTERS_RE.sub(r'', e))
                sheet1.cell(row=count, column=7, value=ILLEGAL_CHARACTERS_RE.sub(r'', f))
                sheet1.cell(row=count, column=8, value=ILLEGAL_CHARACTERS_RE.sub(r'', g))
                sheet1.cell(row=count, column=9, value=ILLEGAL_CHARACTERS_RE.sub(r'', h))
                count += 1

            wb.save(self.input_excel)  # 注意这里是test
            print("EXCEL写入成功!")
        except:
            print("EXCEL写入失败!")
